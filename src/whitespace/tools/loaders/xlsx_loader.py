"""XLSX loader (openpyxl)."""

from __future__ import annotations

import logging
from pathlib import Path

from openpyxl import load_workbook

logger = logging.getLogger(__name__)


class XlsxLoader:
    """Extracts text from an XLSX workbook, sheet by sheet."""

    def load(self, path: str) -> str:
        if not Path(path).is_file():
            raise ValueError(f"XLSX not found: {path}")
        parts: list[str] = []
        try:
            workbook = load_workbook(path, read_only=True, data_only=True)
            try:
                for sheet in workbook.worksheets:
                    parts.append(f"# Sheet: {sheet.title}")
                    for row in sheet.iter_rows(values_only=True):
                        cells = [str(c) for c in row if c is not None and str(c) != ""]
                        if cells:
                            parts.append(" | ".join(cells))
            finally:
                workbook.close()
        except Exception as exc:
            logger.warning("XlsxLoader: best-effort extraction for %s (%s)", path, exc)
        return "\n".join(parts)
